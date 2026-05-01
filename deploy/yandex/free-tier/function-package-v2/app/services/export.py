import csv
import io
import json
from typing import List, Dict, Any
from datetime import datetime
from fastapi.responses import StreamingResponse

class ExportService:
    @staticmethod
    def export_to_csv(data: List[Dict], filename: str) -> StreamingResponse:
        """Export data to CSV"""
        if not data:
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(["No data"])
        else:
            output = io.StringIO()
            writer = csv.DictWriter(output, fieldnames=data[0].keys())
            writer.writeheader()
            writer.writerows(data)
        
        output.seek(0)
        return StreamingResponse(
            output,
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}_{datetime.now().strftime('%Y%m%d')}.csv"}
        )
    
    @staticmethod
    def export_to_excel(data: List[Dict], filename: str) -> StreamingResponse:
        """Export data to Excel"""
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Export"
        
        if data:
            # Headers
            headers = list(data[0].keys())
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Data
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row=row_idx, column=col_idx, value=str(row_data.get(header, "")))
            
            # Auto-size columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}_{datetime.now().strftime('%Y%m%d')}.xlsx"}
        )
    
    @staticmethod
    def export_to_json(data: List[Dict], filename: str) -> StreamingResponse:
        """Export data to JSON"""
        output = io.StringIO()
        json.dump(data, output, ensure_ascii=False, indent=2, default=str)
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/json",
            headers={"Content-Disposition": f"attachment; filename={filename}_{datetime.now().strftime('%Y%m%d')}.json"}
        )
